# uncompyle6 version 3.7.4
# Python bytecode 3.7 (3394)
# Decompiled from: Python 3.6.9 (default, Apr 18 2020, 01:56:04) 
# [GCC 8.4.0]
# Embedded file name: C:\Users\Docode\Desktop\Gimbow\DoCodeCarga\procesos\procesos.py
# Compiled at: 2020-01-08 11:15:30
# Size of source mod 2**32: 7615 bytes
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import docode_managebd.procesos_bd as proc_bd
from datetime import datetime
from django.http import HttpResponse

def leerExcel(excel_file, modelo):
    respuesta = {'resp':1, 
     'mensaje':''}
    campos = proc_bd.obtener_campos(modelo)
    campos.pop(0)
    campos_cont = 1
    dict_cont = 0
    wb = openpyxl.load_workbook(excel_file)
    worksheet = wb['datos']
    for row in worksheet.iter_rows():
        dict_cont = 0
        objeto = modelo()
        for cell in row:
            dato = cell.value
            try:
                if campos_cont >= len(campos):
                    campo_info = campos[dict_cont]
                    if dato == '' or dato == None:
                        dato == None
                    else:
                        if campo_info['tipo'] == 'ForeignKey':
                            querySet = campo_info['qdata']
                            objetoR = querySet.model.objects.filter(id=(int(dato))).first()
                            name = campo_info['nombre']
                            setattr(objeto, name, objetoR)
                            a = 0
                        else:
                            if campo_info['tipo'] == 'DateField':
                                if type(dato) == datetime:
                                    objeto.__dict__[campo_info['nombre']] = dato
                                    a = 0
                            elif type(dato) != datetime:
                                objeto.__dict__[campo_info['nombre']] = obtenerFecha(dato)
                else:
                    if campo_info['tipo'] == 'FileField':
                        pass
                    else:
                        if campo_info['tipo'] != 'DateField':
                            objeto.__dict__[campo_info['nombre']] = dato
                        dict_cont += 1
                    campos_cont += 1
            except Exception as e:
                try:
                    respuesta['resp'] = 0
                    respuesta['mensaje'] = 'Error: ' + str(e)
                finally:
                    e = None
                    del e

        if campos_cont > len(campos):
            objeto.save()

    return response


def obtenerFecha(dato):
    dato = '/'.join(dato.split('-'))
    dato = datetime.strptime(dato, '%d/%m/%Y')


def obtenerLayout(modelo):
    campos = proc_bd.obtener_campos(modelo)
    campos.pop(0)
    sheet_index = 0
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=' + modelo.__name__ + '-{date}.xlsx'.format(date=(datetime.now().strftime('%Y-%m-%d')))
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'datos'
    header_font = Font(name='Calibri', bold=True, color='ffffff')
    centered_alignment = Alignment(horizontal='center')
    border_bottom = Border(bottom=Side(border_style='medium', color='0aa6d4'),
      right=Side(border_style='medium', color='0aa6d4'))
    try:
        columns = []
        for campo in campos:
            if campo['foreign'] == True:
                columns.append(campo['nombre'] + '(id)')
            else:
                if campo['tipo'] == 'FileField':
                    continue
                columns.append(campo['nombre'])

        row_num = 1
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.font = header_font
            cell.border = border_bottom
            cell.alignment = centered_alignment
            cell.fill = PatternFill(start_color='08c7ff', end_color='08c7ff', fill_type='solid')

        for campo in campos:
            if campo['tipo'] == 'choices':
                allChoices = len(campo['qdata'])
                datos = campo['qdata']
                sheet_index += 1
                workbook.create_sheet(campo['nombre'].upper())
                workbook.active = sheet_index
                worksheet = workbook.active
                choices = campo['qdata']
                columns = []
                for choiceID in range(allChoices):
                    if choiceID == 0:
                        columns.append('id')
                    else:
                        columns.append('opciones')
                    row_num = 1

                for col_num, column_title in enumerate(columns, 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.value = column_title
                    cell.font = header_font
                    cell.border = border_bottom
                    cell.alignment = centered_alignment
                    cell.fill = PatternFill(start_color='08c7ff', end_color='08c7ff', fill_type='solid')

                for dato in datos:
                    row_num += 1
                    i = 0
                    for col_num, column_title in enumerate(columns, 1):
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell.value = dato[i]
                        cell.alignment = centered_alignment
                        i += 1

        workbook.save(response)
    except Exception as e:
        try:
            pass
        finally:
            e = None
            del e

    return response