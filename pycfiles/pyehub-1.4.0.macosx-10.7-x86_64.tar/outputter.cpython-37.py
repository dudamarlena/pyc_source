# uncompyle6 version 3.7.4
# Python bytecode 3.7 (3394)
# Decompiled from: Python 3.6.9 (default, Apr 18 2020, 01:56:04) 
# [GCC 8.4.0]
# Embedded file name: /Users/juancomish/miniconda3/lib/python3.7/site-packages/pyehub/outputter.py
# Compiled at: 2019-07-30 13:57:16
# Size of source mod 2**32: 23156 bytes
"""
A script for outputting the results of a PyEHub model.

Outputs can be stored to an excel spreadsheet or printed to the console.
"""
import collections
from collections import OrderedDict
from typing import Union
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
EXCEL_ENGINE = 'openpyxl'
EXCEL_FIRST_COLUMN_WIDTH = 25

def to_dataframes(frames: dict) -> OrderedDict:
    """
    Convert the values of a dictionary into dataframes if they can be converted.

    Args:
        frames: The dictionary to be converted

    Returns:
        An ordered dictionary with dataframes as values (if they can be
        converted).
    """
    for name, value in frames.items():
        if isinstance(value, dict):
            value = to_dataframe(name, value)
        frames[name] = value

    return OrderedDict(sorted(frames.items()))


def to_dataframe(name: str, value: dict) -> pd.DataFrame:
    """
    Convert a dictionary into a dataframe.

    Args:
        name: The name for the dataframe
        value: The dictionary to be converted

    Returns:
        A Pandas DataFrame.
    """
    value = sort_dict(value)
    value = pd.DataFrame.from_dict(value, orient='index')
    if list(value.columns) == [0]:
        value.columns = [
         name]
    num_rows, num_columns = value.shape
    if num_columns > num_rows:
        value = value.T
    return value


def _create_sheet(excel_writer: pd.ExcelWriter, sheet_name: str):
    """
    Create an Excel sheet with the given name.

    Args:
        excel_writer: The ExcelWriter object
        sheet_name: The name of the new sheet

    Returns:
        The worksheet for the new sheet
    """
    pd.DataFrame().to_excel(excel_writer, sheet_name)
    return excel_writer.sheets[sheet_name]


def output_excel(solution: dict, file_name: str, time_steps: int=None, sheets: list=None):
    """ 
    Output the solution dictionary to an Excel file.

    It outputs the time series data in their own sheet with the rest being put
    into another sheet.

    Args:
        solution: A dictionary of the solution part of the response format.
            This contains the variables and parameters of the solved model.
        file_name: The name of the Excel file to write to
        time_steps: Optional. The number of time steps to classify a dataframe
            as holding time series data.
        sheets: A list of all the sheets to be contained in the excel file.
    """
    excel_writer = pd.ExcelWriter(file_name, engine=EXCEL_ENGINE)
    attributes = to_dataframes(solution)

    def _has_time_series_data(value):
        if isinstance(value, pd.DataFrame):
            num_rows, _ = value.shape
            return time_steps and num_rows >= time_steps
        return False

    time_series = {name:value for name, value in attributes.items() if _has_time_series_data(value)}
    non_time_series = {name:value for name, value in attributes.items() if name not in time_series}
    for name, value in time_series.items():
        value.to_excel(excel_writer, f"{name}")

    first_sheet = 'Other' if sheets is None else sheets[0]
    if sheets is not None:
        _dict_to_excel_sheet(excel_writer, first_sheet, non_time_series, padding=1, additional_sheets=(sheets[1:]))
    else:
        _dict_to_excel_sheet(excel_writer, first_sheet, non_time_series, padding=1)
    excel_writer.save()


def _dict_to_excel_sheet(excel_writer, sheet_name, non_time_series, padding=0, additional_sheets: list=None):
    """ 
    Output a dictionary to an excel sheet.

    Args:
        excel_writer: The ExcelWriter object
        sheet_name: The name of the sheet to output to
        non_time_series: The dictionary of data
        padding: The padding between key, value pairs
        additional_sheets: a list containing names of other sheets
                           to potentially be added to the excel document.
    """
    hold_row = 1
    row = 1
    hold_sheet_name = sheet_name
    sheet = _create_sheet(excel_writer, sheet_name)
    hold_sheet = sheet
    for name, value in non_time_series.items():
        if additional_sheets is not None:
            if name in additional_sheets:
                hold_row = row
                row = 1
                sheet = _create_sheet(excel_writer, name)
                sheet_name = name
        if isinstance(value, pd.DataFrame):
            num_rows, num_cols = value.shape
            startrow = max(row - 1, 1)
            value.to_excel(excel_writer, sheet_name, startrow=startrow)
            if num_cols > 1:
                sheet.cell(row=row, column=1).value = name
            row += num_rows + 1
        else:
            sheet.cell(row=row, column=1).value = name
            if not isinstance(value, collections.Iterable):
                value = [value]
            for col, x in enumerate(value, start=2):
                sheet.cell(row=row, column=col).value = str(x)

            row += 1
        if sheet_name != hold_sheet_name:
            sheet_name = hold_sheet_name
            row = hold_row
            sheet = hold_sheet
        row += padding

    sheet.column_dimensions['A'].width = EXCEL_FIRST_COLUMN_WIDTH


def pretty_print(results: dict) -> None:
    """Print the results in a prettier format.

    Args:
        results: The results dictionary to print
    """
    version = results['version']
    solver = results['solver']
    print('Version: {}'.format(version))
    print('Solver')
    print(f"\ttermination condition: {solver['termination_condition']}")
    if solver['termination_condition'] != 'Infeasible':
        print(f"\ttime: {solver['time']}")
        print('Solution')
        print_section('Stuff', results['solution'].copy())


def print_section(section_name: str, solution_section: dict) -> None:
    """
    Print all the attributes with a heading.

    Args:
        section_name: The heading
        solution_section: The dictionary with all the attributes
    """
    half_heading = '=========='
    print(f"\n{half_heading} {section_name} {half_heading}")
    attributes = to_dataframes(solution_section)
    for name, value in attributes.items():
        with pd.option_context('display.max_rows', None, 'display.max_columns', 3):
            print(f"\n{name}: \n{value}")


def sort_dict(mapping: Union[(dict, OrderedDict)]) -> OrderedDict:
    """
    Sorts a dictionary and all its sub-dicionaries as well.

    Examples:
        >>> sort_dict({1: 'a', 3: 'c', 2: 'b'})
        OrderedDict([(1, 'a'), (2, 'b'), (3, 'c')])
        >>> sort_dict({1: {3: 'c', 2: 'b', 1: 'a'}})
        OrderedDict([(1, OrderedDict([(1, 'a'), (2, 'b'), (3, 'c')]))])

    Args:
        mapping: The dictionary to be sorted

    Returns:
        The sorted dictionary as an OrderedDict
    """
    if not isinstance(mapping, dict):
        return mapping
    for key, value in mapping.items():
        mapping[key] = sort_dict(value)

    return OrderedDict(sorted(mapping.items()))


def print_capacities(results):
    """
    Prints the capacities of each tech and storage at the end

    :param results: the solved model

    """
    solution = results['solution']
    capacity_tech = solution['capacity_tech']
    capacity_storage = solution['capacity_storage']
    half_heading = '=========='
    print(f"\n{half_heading} {'Capacities'} {half_heading}")
    print(half_heading * 4)
    print(capacity_storage)
    print(half_heading * 4)
    print(capacity_tech)


def print_warning(results):
    """
    Prints an error if the model burns energy, i.e there is energy from storage and energy to storage
    at same time step.

    :param results: the solved model

    """
    solution = results['solution']
    storages = solution['storages']
    energy_from_storage = solution['energy_from_storage']
    energy_to_storage = solution['energy_to_storage']
    time = len(solution['time'])
    half_heading = '===='
    for storage in storages:
        for t in range(time):
            energy_from = energy_from_storage[storage][t]
            energy_to = energy_to_storage[storage][t]
            if energy_to != 0 and energy_from != 0:
                print(f"{half_heading} {'Warning:'} {half_heading} \n{'Burning energy for storage: '} {storage} {'; For time step: '} {t} ")


def stream_info(results, output_file):
    """
    New output format with the information separated in different sheets for different streams.

    :param results: the solved model
    :param output_file: the output excel file
    """
    solution = results['solution']
    streams = solution['streams']
    load = solution['LOADS']
    energy_from_storage = solution['energy_from_storage']
    energy_to_storage = solution['energy_to_storage']
    energy_exported = solution['energy_exported']
    energy_imported = solution['energy_imported']
    demands = solution['demands']
    time = len(solution['time'])
    export_streams = solution['export_streams']
    import_streams = solution['import_streams']
    storages = solution['storages']
    storage_level = solution['storage_level']
    n = len(storages)
    with pd.ExcelWriter(output_file, engine=EXCEL_ENGINE) as (writer):
        writer.book = load_workbook(output_file)
    keep = ['other', 'TIME_SERIES', 'energy_input', 'capacity_tech', 'capacity_storage']
    for sheet_name in writer.book.sheetnames:
        if sheet_name not in keep:
            del writer.book[sheet_name]

    for stream in streams:
        if stream in demands:
            df = pd.DataFrame({stream + ' Load': [load[stream][t] for t in range(time)]})
            df.to_excel(writer, sheet_name=stream, startrow=(time + 2))
        df1 = pd.DataFrame({'energy_from ' + storage:[energy_from_storage[storage][t] for t in range(time)] for storage in storages})
        df2 = pd.DataFrame({'energy_to ' + storage:[energy_to_storage[storage][t] for t in range(time)] for storage in storages})
        df3 = pd.DataFrame({'level ' + storage:[storage_level[storage][t] for t in range(time)] for storage in storages})
        df1.to_excel(writer, sheet_name=stream)
        df2.to_excel(writer, sheet_name=stream, startcol=(n + 2))
        df3.to_excel(writer, sheet_name=stream, startcol=(2 * (n + 2)))
        if stream in import_streams:
            df4 = pd.DataFrame({stream + ' imported': [energy_imported[stream][t] for t in range(time)]})
            df4.to_excel(writer, sheet_name=stream, startrow=(time + 2), startcol=4)
        if stream in export_streams:
            df5 = pd.DataFrame({stream + ' exported': [energy_exported[stream][t] for t in range(time)]})
            df5.to_excel(writer, sheet_name=stream, startrow=(time + 2), startcol=7)
        writer.save()


def plot_storages--- This code section failed: ---

 L. 382         0  LOAD_FAST                'results'
                2  LOAD_STR                 'solution'
                4  BINARY_SUBSCR    
                6  LOAD_METHOD              copy
                8  CALL_METHOD_0         0  '0 positional arguments'
               10  STORE_FAST               'solution_section'

 L. 383        12  LOAD_GLOBAL              to_dataframes
               14  LOAD_FAST                'solution_section'
               16  CALL_FUNCTION_1       1  '1 positional argument'
               18  STORE_FAST               'attributes'

 L. 385        20  LOAD_FAST                'kwargs'
               22  LOAD_METHOD              get
               24  LOAD_STR                 'pl_state'
               26  LOAD_CONST               True
               28  CALL_METHOD_2         2  '2 positional arguments'
               30  STORE_FAST               'pl_state'

 L. 386        32  LOAD_FAST                'kwargs'
               34  LOAD_METHOD              get
               36  LOAD_STR                 'pl_gross_ch'
               38  LOAD_CONST               True
               40  CALL_METHOD_2         2  '2 positional arguments'
               42  STORE_FAST               'pl_gross_ch'

 L. 387        44  LOAD_FAST                'kwargs'
               46  LOAD_METHOD              get
               48  LOAD_STR                 'pl_gross_dch'
               50  LOAD_CONST               True
               52  CALL_METHOD_2         2  '2 positional arguments'
               54  STORE_FAST               'pl_gross_dch'

 L. 388        56  LOAD_FAST                'kwargs'
               58  LOAD_METHOD              get
               60  LOAD_STR                 'pl_net_ch'
               62  LOAD_CONST               True
               64  CALL_METHOD_2         2  '2 positional arguments'
               66  STORE_FAST               'pl_net_ch'

 L. 389        68  LOAD_FAST                'kwargs'
               70  LOAD_METHOD              get
               72  LOAD_STR                 'pl_net_dch'
               74  LOAD_CONST               True
               76  CALL_METHOD_2         2  '2 positional arguments'
               78  STORE_FAST               'pl_net_dch'

 L. 390        80  LOAD_FAST                'kwargs'
               82  LOAD_METHOD              get
               84  LOAD_STR                 'pl_decay'
               86  LOAD_CONST               True
               88  CALL_METHOD_2         2  '2 positional arguments'
               90  STORE_FAST               'pl_decay'

 L. 391        92  LOAD_FAST                'kwargs'
               94  LOAD_METHOD              get
               96  LOAD_STR                 'size'
               98  LOAD_CONST               (10, 5)
              100  CALL_METHOD_2         2  '2 positional arguments'
              102  STORE_FAST               'size'

 L. 392       104  LOAD_FAST                'size'
              106  LOAD_CONST               0
              108  BINARY_SUBSCR    
              110  LOAD_CONST               0
              112  COMPARE_OP               ==
              114  POP_JUMP_IF_TRUE    128  'to 128'
              116  LOAD_FAST                'size'
              118  LOAD_CONST               1
              120  BINARY_SUBSCR    
              122  LOAD_CONST               0
              124  COMPARE_OP               ==
              126  POP_JUMP_IF_FALSE   136  'to 136'
            128_0  COME_FROM           114  '114'

 L. 393       128  LOAD_GLOBAL              ValueError
              130  LOAD_STR                 "Please pass non-zero values in the 'size' tuple."
              132  CALL_FUNCTION_1       1  '1 positional argument'
              134  RAISE_VARARGS_1       1  'exception instance'
            136_0  COME_FROM           126  '126'

 L. 394       136  LOAD_FAST                'kwargs'
              138  LOAD_METHOD              get
              140  LOAD_STR                 'percentage'
              142  LOAD_CONST               False
              144  CALL_METHOD_2         2  '2 positional arguments'
              146  STORE_FAST               'percentage'

 L. 396   148_150  SETUP_LOOP          834  'to 834'
              152  LOAD_FAST                'attributes'
              154  LOAD_STR                 'storages'
              156  BINARY_SUBSCR    
              158  GET_ITER         
            160_0  COME_FROM           768  '768'
          160_162  FOR_ITER            832  'to 832'
              164  STORE_FAST               'storage'

 L. 397       166  LOAD_GLOBAL              plt
              168  LOAD_ATTR                subplots
              170  LOAD_CONST               1
              172  LOAD_CONST               2
              174  LOAD_CONST               True
              176  LOAD_FAST                'size'
              178  LOAD_CONST               ('nrows', 'ncols', 'sharey', 'figsize')
              180  CALL_FUNCTION_KW_4     4  '4 total positional and keyword args'
              182  UNPACK_SEQUENCE_2     2 
              184  STORE_FAST               'fig'
              186  STORE_FAST               'axes'

 L. 398       188  LOAD_FAST                'fig'
              190  LOAD_ATTR                text
              192  LOAD_CONST               0.524
              194  LOAD_CONST               0.02
              196  LOAD_STR                 'Timesteps (-)'
              198  LOAD_STR                 'center'
              200  LOAD_STR                 'center'
              202  LOAD_CONST               ('ha', 'va')
              204  CALL_FUNCTION_KW_5     5  '5 total positional and keyword args'
              206  POP_TOP          

 L. 399       208  LOAD_FAST                'percentage'
              210  POP_JUMP_IF_FALSE   236  'to 236'

 L. 400       212  LOAD_FAST                'fig'
              214  LOAD_ATTR                text
              216  LOAD_CONST               0.01
              218  LOAD_CONST               0.5
              220  LOAD_STR                 'Storage state [%]'
              222  LOAD_STR                 'center'
              224  LOAD_STR                 'center'
              226  LOAD_STR                 'vertical'
              228  LOAD_CONST               ('ha', 'va', 'rotation')
              230  CALL_FUNCTION_KW_6     6  '6 total positional and keyword args'
              232  POP_TOP          
              234  JUMP_FORWARD        258  'to 258'
            236_0  COME_FROM           210  '210'

 L. 402       236  LOAD_FAST                'fig'
              238  LOAD_ATTR                text
              240  LOAD_CONST               0.01
              242  LOAD_CONST               0.5
              244  LOAD_STR                 'Storage state (kWh)'
              246  LOAD_STR                 'center'
              248  LOAD_STR                 'center'
              250  LOAD_STR                 'vertical'
              252  LOAD_CONST               ('ha', 'va', 'rotation')
              254  CALL_FUNCTION_KW_6     6  '6 total positional and keyword args'
              256  POP_TOP          
            258_0  COME_FROM           234  '234'

 L. 403       258  LOAD_FAST                'fig'
              260  LOAD_ATTR                text
              262  LOAD_CONST               0.5
              264  LOAD_CONST               0.97
              266  LOAD_FAST                'storage'
              268  LOAD_STR                 'center'
              270  LOAD_STR                 'center'
              272  LOAD_CONST               ('ha', 'va')
              274  CALL_FUNCTION_KW_5     5  '5 total positional and keyword args'
              276  POP_TOP          

 L. 404       278  LOAD_FAST                'axes'
              280  LOAD_CONST               0
              282  BINARY_SUBSCR    
              284  STORE_FAST               'ax'

 L. 405       286  LOAD_FAST                'axes'
              288  LOAD_CONST               1
              290  BINARY_SUBSCR    
              292  STORE_FAST               'ax1'

 L. 408       294  LOAD_FAST                'attributes'
              296  LOAD_STR                 'storage_level'
              298  BINARY_SUBSCR    
              300  LOAD_FAST                'storage'
              302  BINARY_SUBSCR    
              304  STORE_FAST               'ser_storage_state'

 L. 409       306  LOAD_FAST                'attributes'
              308  LOAD_STR                 'energy_to_storage'
              310  BINARY_SUBSCR    
              312  LOAD_FAST                'storage'
              314  BINARY_SUBSCR    
              316  STORE_FAST               'ser_gross_charge'

 L. 410       318  LOAD_FAST                'attributes'
              320  LOAD_STR                 'energy_from_storage'
              322  BINARY_SUBSCR    
              324  LOAD_FAST                'storage'
              326  BINARY_SUBSCR    
              328  UNARY_NEGATIVE   
              330  STORE_FAST               'ser_gross_discharge'

 L. 412       332  LOAD_FAST                'attributes'
              334  LOAD_STR                 'CHARGING_EFFICIENCY'
              336  BINARY_SUBSCR    
              338  LOAD_STR                 'CHARGING_EFFICIENCY'
              340  BINARY_SUBSCR    
              342  LOAD_FAST                'storage'
              344  BINARY_SUBSCR    
              346  STORE_FAST               'eff_charge'

 L. 413       348  LOAD_FAST                'attributes'
              350  LOAD_STR                 'DISCHARGING_EFFICIENCY'
              352  BINARY_SUBSCR    
              354  LOAD_STR                 'DISCHARGING_EFFICIENCY'
              356  BINARY_SUBSCR    
              358  LOAD_FAST                'storage'
              360  BINARY_SUBSCR    
              362  STORE_FAST               'eff_discharge'

 L. 414       364  LOAD_FAST                'attributes'
              366  LOAD_STR                 'STORAGE_STANDING_LOSSES'
              368  BINARY_SUBSCR    
              370  LOAD_STR                 'STORAGE_STANDING_LOSSES'
              372  BINARY_SUBSCR    
              374  LOAD_FAST                'storage'
              376  BINARY_SUBSCR    
              378  STORE_FAST               'decay'

 L. 415       380  LOAD_FAST                'attributes'
              382  LOAD_STR                 'capacity_storage'
              384  BINARY_SUBSCR    
              386  LOAD_STR                 'capacity_storage'
              388  BINARY_SUBSCR    
              390  LOAD_FAST                'storage'
              392  BINARY_SUBSCR    
              394  STORE_FAST               'capacity'

 L. 417       396  LOAD_CONST               0
              398  STORE_FAST               'ser_net_discharge'

 L. 418       400  LOAD_FAST                'ser_gross_charge'
              402  LOAD_METHOD              multiply
              404  LOAD_FAST                'eff_charge'
              406  CALL_METHOD_1         1  '1 positional argument'
              408  STORE_FAST               'ser_net_charge'

 L. 419       410  LOAD_FAST                'eff_discharge'
              412  LOAD_CONST               0
              414  COMPARE_OP               !=
          416_418  POP_JUMP_IF_FALSE   430  'to 430'

 L. 420       420  LOAD_FAST                'ser_gross_discharge'
              422  LOAD_METHOD              divide
              424  LOAD_FAST                'eff_discharge'
              426  CALL_METHOD_1         1  '1 positional argument'
              428  STORE_FAST               'ser_net_discharge'
            430_0  COME_FROM           416  '416'

 L. 421       430  LOAD_FAST                'ser_storage_state'
              432  LOAD_METHOD              multiply
              434  LOAD_FAST                'decay'
              436  CALL_METHOD_1         1  '1 positional argument'
              438  LOAD_CONST               None
              440  LOAD_CONST               -1
              442  BUILD_SLICE_2         2 
              444  BINARY_SUBSCR    
              446  UNARY_NEGATIVE   
              448  STORE_FAST               'decay_loss'

 L. 423       450  LOAD_FAST                'ser_storage_state'
              452  LOAD_FAST                'ser_gross_charge'
              454  LOAD_FAST                'ser_net_charge'

 L. 424       456  LOAD_FAST                'ser_gross_discharge'
              458  LOAD_FAST                'ser_net_discharge'
              460  LOAD_FAST                'decay_loss'
              462  LOAD_CONST               ('storage state', 'charge from stream', 'net charge', 'discharge to stream', 'net discharge', 'standing loss')
              464  BUILD_CONST_KEY_MAP_6     6 
              466  STORE_FAST               'd'

 L. 425       468  LOAD_GLOBAL              pd
              470  LOAD_ATTR                DataFrame
              472  LOAD_FAST                'd'
              474  LOAD_CONST               ('data',)
              476  CALL_FUNCTION_KW_1     1  '1 total positional and keyword args'
              478  STORE_FAST               'df'

 L. 427       480  LOAD_FAST                'percentage'
          482_484  POP_JUMP_IF_FALSE   512  'to 512'

 L. 428       486  LOAD_FAST                'capacity'
              488  LOAD_CONST               0
              490  COMPARE_OP               !=
          492_494  POP_JUMP_IF_FALSE   512  'to 512'

 L. 429       496  LOAD_FAST                'df'
              498  LOAD_METHOD              div
              500  LOAD_FAST                'capacity'
              502  CALL_METHOD_1         1  '1 positional argument'
              504  LOAD_METHOD              mul
              506  LOAD_CONST               100
              508  CALL_METHOD_1         1  '1 positional argument'
              510  STORE_FAST               'df'
            512_0  COME_FROM           492  '492'
            512_1  COME_FROM           482  '482'

 L. 431       512  LOAD_FAST                'pl_state'
          514_516  POP_JUMP_IF_FALSE   542  'to 542'

 L. 432       518  LOAD_FAST                'df'
              520  LOAD_STR                 'storage state'
              522  BINARY_SUBSCR    
              524  LOAD_ATTR                plot
              526  LOAD_STR                 'bar'
              528  LOAD_STR                 ' '
              530  LOAD_FAST                'ax'
              532  LOAD_STR                 'deepskyblue'
              534  LOAD_CONST               True
              536  LOAD_CONST               ('kind', 'title', 'ax', 'color', 'legend')
              538  CALL_FUNCTION_KW_5     5  '5 total positional and keyword args'
              540  POP_TOP          
            542_0  COME_FROM           514  '514'

 L. 433       542  LOAD_FAST                'pl_gross_ch'
          544_546  POP_JUMP_IF_FALSE   574  'to 574'

 L. 434       548  LOAD_FAST                'df'
              550  LOAD_STR                 'charge from stream'
              552  BINARY_SUBSCR    
              554  LOAD_ATTR                plot
              556  LOAD_STR                 'steps-post'
              558  LOAD_STR                 ' '
              560  LOAD_STR                 'lightgreen'

 L. 435       562  LOAD_CONST               3
              564  LOAD_FAST                'ax1'
              566  LOAD_CONST               True
              568  LOAD_CONST               ('drawstyle', 'title', 'color', 'linewidth', 'ax', 'legend')
              570  CALL_FUNCTION_KW_6     6  '6 total positional and keyword args'
              572  POP_TOP          
            574_0  COME_FROM           544  '544'

 L. 436       574  LOAD_FAST                'pl_net_ch'
          576_578  POP_JUMP_IF_FALSE   608  'to 608'

 L. 437       580  LOAD_FAST                'df'
              582  LOAD_STR                 'net charge'
              584  BINARY_SUBSCR    
              586  LOAD_ATTR                plot
              588  LOAD_STR                 'steps-post'
              590  LOAD_STR                 ' '
              592  LOAD_STR                 'green'

 L. 438       594  LOAD_STR                 '--'
              596  LOAD_CONST               2
              598  LOAD_FAST                'ax1'
              600  LOAD_CONST               True
              602  LOAD_CONST               ('drawstyle', 'title', 'color', 'linestyle', 'linewidth', 'ax', 'legend')
              604  CALL_FUNCTION_KW_7     7  '7 total positional and keyword args'
              606  POP_TOP          
            608_0  COME_FROM           576  '576'

 L. 439       608  LOAD_FAST                'pl_gross_dch'
          610_612  POP_JUMP_IF_FALSE   640  'to 640'

 L. 440       614  LOAD_FAST                'df'
              616  LOAD_STR                 'discharge to stream'
              618  BINARY_SUBSCR    
              620  LOAD_ATTR                plot
              622  LOAD_STR                 'steps-post'
              624  LOAD_STR                 ' '
              626  LOAD_STR                 'orange'

 L. 441       628  LOAD_CONST               3
              630  LOAD_FAST                'ax1'
              632  LOAD_CONST               True
              634  LOAD_CONST               ('drawstyle', 'title', 'color', 'linewidth', 'ax', 'legend')
              636  CALL_FUNCTION_KW_6     6  '6 total positional and keyword args'
              638  POP_TOP          
            640_0  COME_FROM           610  '610'

 L. 442       640  LOAD_FAST                'pl_net_dch'
          642_644  POP_JUMP_IF_FALSE   674  'to 674'

 L. 443       646  LOAD_FAST                'df'
              648  LOAD_STR                 'net discharge'
              650  BINARY_SUBSCR    
              652  LOAD_ATTR                plot
              654  LOAD_STR                 'steps-post'
              656  LOAD_STR                 ' '
              658  LOAD_STR                 'red'

 L. 444       660  LOAD_STR                 '--'
              662  LOAD_CONST               2
              664  LOAD_FAST                'ax1'
              666  LOAD_CONST               True
              668  LOAD_CONST               ('drawstyle', 'title', 'color', 'linestyle', 'linewidth', 'ax', 'legend')
              670  CALL_FUNCTION_KW_7     7  '7 total positional and keyword args'
              672  POP_TOP          
            674_0  COME_FROM           642  '642'

 L. 445       674  LOAD_FAST                'pl_decay'
          676_678  POP_JUMP_IF_FALSE   704  'to 704'

 L. 446       680  LOAD_FAST                'df'
              682  LOAD_STR                 'standing loss'
              684  BINARY_SUBSCR    
              686  LOAD_ATTR                plot
              688  LOAD_STR                 'bar'
              690  LOAD_STR                 ' '
              692  LOAD_STR                 'saddlebrown'

 L. 447       694  LOAD_FAST                'ax'
              696  LOAD_CONST               True
              698  LOAD_CONST               ('kind', 'title', 'color', 'ax', 'legend')
              700  CALL_FUNCTION_KW_5     5  '5 total positional and keyword args'
              702  POP_TOP          
            704_0  COME_FROM           676  '676'

 L. 448       704  LOAD_FAST                'ax1'
              706  LOAD_METHOD              set_xticks
              708  LOAD_FAST                'df'
              710  LOAD_ATTR                index
              712  LOAD_CONST               None
              714  LOAD_CONST               -1
              716  BUILD_SLICE_2         2 
              718  BINARY_SUBSCR    
              720  CALL_METHOD_1         1  '1 positional argument'
              722  POP_TOP          

 L. 449       724  LOAD_FAST                'ax'
              726  LOAD_METHOD              set_xticks
              728  LOAD_FAST                'df'
              730  LOAD_ATTR                index
              732  CALL_METHOD_1         1  '1 positional argument'
              734  POP_TOP          

 L. 452       736  LOAD_FAST                'pl_state'
          738_740  POP_JUMP_IF_TRUE    770  'to 770'
              742  LOAD_FAST                'pl_gross_ch'
          744_746  POP_JUMP_IF_TRUE    770  'to 770'
              748  LOAD_FAST                'pl_net_ch'
          750_752  POP_JUMP_IF_TRUE    770  'to 770'
              754  LOAD_FAST                'pl_gross_dch'
          756_758  POP_JUMP_IF_TRUE    770  'to 770'

 L. 453       760  LOAD_FAST                'pl_net_dch'
          762_764  POP_JUMP_IF_TRUE    770  'to 770'
              766  LOAD_FAST                'pl_decay'
              768  POP_JUMP_IF_FALSE   160  'to 160'
            770_0  COME_FROM           762  '762'
            770_1  COME_FROM           756  '756'
            770_2  COME_FROM           750  '750'
            770_3  COME_FROM           744  '744'
            770_4  COME_FROM           738  '738'

 L. 454       770  LOAD_GLOBAL              plt
              772  LOAD_ATTR                legend
              774  LOAD_STR                 'best'
              776  LOAD_CONST               ('loc',)
              778  CALL_FUNCTION_KW_1     1  '1 total positional and keyword args'
              780  POP_TOP          

 L. 455       782  SETUP_LOOP          822  'to 822'
              784  LOAD_FAST                'fig'
              786  LOAD_ATTR                axes
              788  GET_ITER         
              790  FOR_ITER            820  'to 820'
              792  STORE_FAST               'ax'

 L. 456       794  LOAD_GLOBAL              plt
              796  LOAD_METHOD              sca
              798  LOAD_FAST                'ax'
              800  CALL_METHOD_1         1  '1 positional argument'
              802  POP_TOP          

 L. 457       804  LOAD_GLOBAL              plt
              806  LOAD_ATTR                xticks
              808  LOAD_CONST               0
              810  LOAD_CONST               ('rotation',)
              812  CALL_FUNCTION_KW_1     1  '1 total positional and keyword args'
              814  POP_TOP          
          816_818  JUMP_BACK           790  'to 790'
              820  POP_BLOCK        
            822_0  COME_FROM_LOOP      782  '782'

 L. 458       822  LOAD_GLOBAL              plt
              824  LOAD_METHOD              show
              826  CALL_METHOD_0         0  '0 positional arguments'
              828  POP_TOP          
              830  JUMP_BACK           160  'to 160'
              832  POP_BLOCK        
            834_0  COME_FROM_LOOP      148  '148'

Parse error at or near `POP_BLOCK' instruction at offset 832


def plot_energy_balance(model, results: dict, **kwargs) -> None:
    """
    Visualization of energy balance for all the streams, i.e., Plots the energy interactions of all the streams with
    loads, converters, storages, imports and exports. Plots are dashed and marked with shape markers for visibility
    in case of overlap. Adjust the properties of the plot by passing arguments from below.

    Args:
        model: The object of EHubModel class (or any child class thereof); the energy hub model.
        results: dictionary; returned by solve() method.
        size: tuple; (width, height) of the plot [default is (9,5)].
        lw: float; linewidth of the plots [default is 2].
        dl: float; length of the dashes constituting the dashed plots [default is 3].
    """
    attributes = results['solution'].copy()
    streams_wo_sources = [x for x in model.streams if x not in model.sources]
    size = kwargs.get('size', (9, 5))
    if size[0] == 0 or size[1] == 0:
        raise ValueError("Please pass non-zero values in the 'size' tuple.")
    lw = kwargs.get('lw', 2)
    dl = kwargs.get('dl', 3)
    for stream in streams_wo_sources:
        fig, ax = plt.subplots(nrows=1, ncols=1, figsize=size)
        dict_data_pos = {}
        dict_data_neg = {}
        for t in model.time:
            curr_neg = curr_pos = 0
            if stream in model.demands:
                load = model.LOADS[stream][t]
                curr_neg += -load
                place_in_dict(t, model, dict_data_neg, 'LOAD', curr_neg)
            for storage in model._get_storages_from_stream(stream):
                q_out = float(attributes['energy_to_storage'][t][storage.name])
                curr_neg += -q_out
                q_in = float(attributes['energy_from_storage'][t][storage.name])
                curr_pos += q_in
                place_in_dict(t, model, dict_data_neg, 'To ' + storage.name + ' (S)', curr_neg)
                place_in_dict(t, model, dict_data_pos, 'From ' + storage.name + ' (S)', curr_pos)

            for tech in model.technologies:
                conversion_rate = float(attributes['CONVERSION_EFFICIENCY'][tech][stream])
                if conversion_rate < 0:
                    energy_input = float(attributes['energy_input'][t][tech])
                    curr_neg += -energy_input
                    place_in_dict(t, model, dict_data_neg, 'To ' + tech + ' (C)', curr_neg)
                if conversion_rate > 0:
                    energy_input = float(attributes['energy_input'][t][tech])
                    energy_output = energy_input * conversion_rate
                    curr_pos += energy_output
                    place_in_dict(t, model, dict_data_pos, 'From ' + tech + ' (C)', curr_pos)

            if stream in model.export_streams:
                energy_exported = attributes['energy_exported'][t][stream]
                curr_neg += -energy_exported
                place_in_dict(t, model, dict_data_neg, 'Exported', curr_neg)
            if stream in model.import_streams:
                energy_imported = attributes['energy_imported'][t][stream]
                curr_pos += energy_imported
                place_in_dict(t, model, dict_data_pos, 'Imported', curr_pos)

        x = 1
        y = 1
        mk_c = [
         '1', '2', '3', '4']
        mk_s = ['*', 'P']
        mk_index_c = 0
        mk_index_s = 0
        for label, data in dict_data_neg.items():
            if label.endswith('LOAD'):
                ds = ()
                mk = 's'
                ms = 7
            if label.endswith('(C)'):
                ds = (
                 dl, x)
                x += 1
                mk = mk_c[(mk_index_c % 4)]
                mk_index_c += 1
                ms = 12
            if label.endswith('(S)'):
                ds = (
                 dl, y, 1, y)
                y += 1
                mk = mk_s[(mk_index_s % 2)]
                mk_index_s += 1
                ms = 7.5
            if label.endswith('Exported'):
                mk = '|'
                ds = (dl, 1, 1.5, 1, 1.5, 1, 1.5, 1, 1.5, 1, 1.5, 1)
                ms = 12
            ax.plot(data, label=label, linewidth=lw, dashes=ds, marker=mk, markersize=ms)

        handle_demand, label_demand = ax.get_legend_handles_labels()
        x = 1.2
        y = 1.2
        mk_c = [
         '1', '2', '3', '4']
        mk_s = ['+', 'x']
        mk_index_c = 0
        mk_index_s = 0
        for label, data in dict_data_pos.items():
            if label.endswith('(C)'):
                ds = (
                 dl, x)
                x += 1
                mk = mk_c[(mk_index_c % 4)]
                mk_index_c += 1
                ms = 12
            if label.endswith('(S)'):
                ds = (
                 dl, y, 1, y)
                y += 1
                mk = mk_s[(mk_index_s % 2)]
                mk_index_s += 1
                ms = 12
            if label.endswith('Imported'):
                ds = (
                 dl, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1)
                mk = '|'
                ms = 12
            ax.plot(data, label=label, linewidth=lw, dashes=ds, marker=mk, markersize=ms)

        ax.set(title=stream, xticks=(model.time), xlabel='Timesteps (-)', ylabel=(stream + ' (kWh)'))
        blank_handle = [
         plt.plot([], marker='', ls='')[0]]
        handle_supply, label_supply = ax.get_legend_handles_labels()
        handle_supply = handle_supply[len(handle_demand):]
        label_supply = label_supply[len(label_demand):]
        handles = blank_handle + handle_demand + blank_handle * 2 + handle_supply
        labels = [
         '$\\bf{Demands}$'] + label_demand + [' '] + ['$\\bf{Supplies}$'] + label_supply
        plt.legend(loc='best', bbox_to_anchor=(1, 1), labels=labels, handles=handles)
        plt.show()


def place_in_dict(t, model, dictionary: dict, key: str, value):
    """Appends an entry in the list corresponding to key if key exists in dictionary.
    Otherwise creates a single entry in the dictionary as {key: [value]}"""
    if t == model.time[0]:
        dictionary.update({key: [value]})
    else:
        dictionary[key].append(value)