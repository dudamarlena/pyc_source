# uncompyle6 version 3.7.4
# Python bytecode 3.7 (3394)
# Decompiled from: Python 3.6.9 (default, Apr 18 2020, 01:56:04) 
# [GCC 8.4.0]
# Embedded file name: /Users/juancomish/miniconda3/lib/python3.7/site-packages/pyehub/BatchRunExcel.py
# Compiled at: 2019-07-04 16:16:54
# Size of source mod 2**32: 18440 bytes
"""
A script used to solve a batch series of energy hubs in a combined PyEHub model.

This functionality is now also provided by the multiple hubs module.
"""
import argparse, os, re, shutil, openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment
import multiple_hubs

def get_output_fields(filenames: list, directory: str=None):
    """ get_output_fields(list, str)

                Gets the required outputs from the excel files passed in the
                list of filenames.

                Args:
                        filenames(list): A list of the excel files to find outputs for.

                        directory(str): The optional dicretory that these filenames are located at.
                                        for use only if the filenames do not include their directory.

                Returns:
                        A list with every output requested from every file (duplicates are deleted.)
        """
    master_field_list = []
    for filename in filenames:
        if directory is not None:
            book = openpyxl.load_workbook(directory + filename)
        else:
            book = openpyxl.load_workbook(filename)
        sheet = book['Capacities']
        fields = sheet.iter_rows(min_row=1, max_row=1, min_col=2)
        for cell in list(fields)[0]:
            master_field_list.append(cell.value)

    master_field_list.append('total_cost')
    master_field_list.append('total_carbon')
    return_list = []
    for item in master_field_list:
        if item is not None:
            return_list.append(item)

    return sorted(list(set(return_list)))


def output_to_excel(outputs: dict, sheet, output_dict: dict):
    """ output_to_excel(dict, openpyxl.sheet, dict)

                Takes the output dictionary produced by multiple hubs, and
                writes specific fields to the passed excel sheet based on
                the output dictionary passed.

                Args:
                        outputs(dict): The dictionary produced by multiple_hubs.

                        sheet(openpyxl.sheet): The excel sheet to be written to.

                        output_dict(dict): A dictionary that links keys found in the
                                           outputs dict, and the column that they should
                                           appear in, within the sheet.

        """
    cols = list(sheet.iter_cols(min_col=1, max_col=1))[0]
    first_empty_row = len(cols) + 1
    sheet.merge_cells('A{}:B{}'.format(first_empty_row, first_empty_row))
    sheet.merge_cells('C{}:D{}'.format(first_empty_row, first_empty_row))
    sheet.merge_cells('E{}:F{}'.format(first_empty_row, first_empty_row))
    sheet.merge_cells('G{}:H{}'.format(first_empty_row, first_empty_row))
    for key, value in output_dict.items():
        try:
            sheet['{}{}'.format(value, first_empty_row)] = outputs[key]
        except:
            continue


def main(batch_run_location: str='excel_files/batch_input/', output_directory: str='batchrun_outputs/', epsilon_n: int=0):
    if output_directory is None:
        output_directory = 'batchrun_outputs/'
    else:
        if output_directory[(-1)] != '/':
            if output_directory[(-1)] != '\\':
                output_directory = output_directory + '/'
        else:
            if batch_run_location is None:
                batch_run_location = 'excel_files/batch_input/'
            else:
                if batch_run_location[(-1)] != '/':
                    if batch_run_location[(-1)] != '\\':
                        batch_run_location = batch_run_location + '/'
            if epsilon_n is None:
                epsilon_n = 0
            os.path.exists(output_directory) or os.makedirs(output_directory)
        excel_files = [file for file in os.listdir(batch_run_location) if file[-5:] == '.xlsx' if file[0] != '$']
        file_data = []
        for file in excel_files:
            _position = 0
            for character in file[::-1]:
                if character == '_':
                    break
                _position -= 1

            file_data.append((file, file[_position:-5]))

        file_data = sorted(file_data, key=(lambda x: x[1]))
        file_sets = []
        for index, file in enumerate(file_data[::2]):
            file_sets.append((file, file_data[(index * 2 + 1)]))

        all_fields = get_output_fields(excel_files, directory=batch_run_location)
        first_letter = 'I'
        extra_fields = 2
        final_letter = str(first_letter)
        for _ in range(len(all_fields) - 1 - extra_fields):
            if final_letter[(-1)] != 'Z':
                final_letter = final_letter[:-1] + chr(ord(final_letter[(-1)]) + 1)
            else:
                final_letter = final_letter[:-1] + 'AA'

        extended = [final_letter]
        for _ in range(extra_fields):
            if extended[(-1)][(-1)] != 'Z':
                extended.append(extended[(-1)][:-1] + chr(ord(extended[(-1)][(-1)]) + 1))
            else:
                extended.append(extended[(-1)][:-1] + 'AA')

        output_book = Workbook()
        sheet = output_book.active
        sheet.merge_cells('A1:B2')
        sheet.merge_cells('C1:D2')
        sheet.merge_cells('E1:F2')
        sheet.merge_cells('G1:H2')
        sheet.merge_cells('{}1:{}1'.format(first_letter, final_letter))
        sheet.merge_cells('{}1:{}1'.format(extended[1], extended[2]))
        sheet['A1'] = 'File Name'
        sheet['C1'] = 'Function'
        sheet['E1'] = 'fjob'
        sheet['G1'] = 'Hub'
        sheet['{}1'.format(first_letter)] = 'Capacity'
        sheet['{}1'.format(extended[1])] = 'Totals'
        pos_dict = {'Filename':'A', 
         'Function':'C', 
         'fjob':'E', 
         'Hub':'G'}
        fletter = str(first_letter)
        for field in all_fields:
            sheet['{}2'.format(fletter)] = field
            pos_dict[field] = fletter
            if fletter[(-1)] != 'Z':
                fletter = fletter[:-1] + chr(ord(fletter[(-1)]) + 1)
            else:
                fletter = fletter[:-1] + 'AA'

        sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
        sheet['C1'].alignment = Alignment(horizontal='center', vertical='center')
        sheet['E1'].alignment = Alignment(horizontal='center', vertical='center')
        sheet['G1'].alignment = Alignment(horizontal='center', vertical='center')
        sheet['{}1'.format(first_letter)].alignment = Alignment(horizontal='center', vertical='center')
        sheet['{}1'.format(extended[1])].alignment = Alignment(horizontal='center', vertical='center')
        for fjob, file_set in enumerate(file_sets):
            shutil.copyfile(batch_run_location + file_set[0][0], 'docs/tutorials/network/hub1.xlsx')
            shutil.copyfile(batch_run_location + file_set[1][0], 'docs/tutorials/network/hub2.xlsx')
            print('Finding minimum cost for {} and {}'.format(file_set[0][0][:-5], file_set[1][0][:-5]))
            cost_min = multiple_hubs(False, output_directory + file_set[0][0][:_position - 1])
            hub1_output_fields = get_output_fields([file_set[0][0]], directory=batch_run_location)
            hub2_output_fields = get_output_fields([file_set[1][0]], directory=batch_run_location)
            hub1_outputs = {'Filename':file_set[0][0][:-5], 
             'Function':'Minimize Cost', 
             'fjob':2 * fjob + 1, 
             'Hub':1}
            hub2_outputs = {'Filename':file_set[1][0][:-5], 
             'Function':'Minimize Cost', 
             'fjob':2 * fjob + 1, 
             'Hub':2}
            for field in hub1_output_fields:
                hub1_outputs[field] = cost_min[0][field]

            for field in hub2_output_fields:
                hub2_outputs[field] = cost_min[1][field]

            output_to_excel(outputs=hub1_outputs, sheet=sheet, output_dict=pos_dict)
            output_to_excel(outputs=hub2_outputs, sheet=sheet, output_dict=pos_dict)
            avg_carbon = (cost_min[0]['total_carbon'] + cost_min[1]['total_carbon']) / 2
            carbon_per_step = 0 if epsilon_n == 0 else avg_carbon / epsilon_n
            for n in range(epsilon_n + 1):
                print('Running {} and {}  at n = {}, max carbon = {}'.format(file_set[0][0][:-5], file_set[1][0][:-5], n, carbon_per_step * n))
                carbon_min = multiple_hubs(True, output_directory + 'n-{}'.format(n) + file_set[0][0][:_position - 1], carbon_per_step * n)
                hub1_outputs = {'Filename':file_set[0][0][:-5], 
                 'Function':'Minimize Carbon; n = {}'.format(n), 
                 'fjob':fjob + 1 * 2, 
                 'Hub':1}
                hub2_outputs = {'Filename':file_set[1][0][:-5], 
                 'Function':'Minimize Carbon; n = {}'.format(n), 
                 'fjob':fjob + 1 * 2, 
                 'Hub':2}
                for field in hub1_output_fields:
                    hub1_outputs[field] = carbon_min[0][field]

                for field in hub2_output_fields:
                    hub2_outputs[field] = carbon_min[1][field]

                output_to_excel(outputs=hub1_outputs, sheet=sheet, output_dict=pos_dict)
                output_to_excel(outputs=hub2_outputs, sheet=sheet, output_dict=pos_dict)

        output_book.save(output_directory + 'final_results.xlsx')


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='A commandline utility to batch run tests.')
    parser.add_argument('-d', '--directory_name', type=str, help='The directory containing the excel file to be run.\nDefault is excel_files/batch_input/')
    parser.add_argument('-o', '--output_name', type=str, help='The directory to store all outputted files in..\nDefault is batchrun_outputs/')
    parser.add_argument('-n', '--epsilon_n', type=int, help='The number of steps to break into for the epsilon constraint.')
    args = parser.parse_args()
    main(args.directory_name, args.output_name, args.epsilon_n)