# uncompyle6 version 3.7.4
# Python bytecode 3.6 (3379)
# Decompiled from: Python 3.6.9 (default, Apr 18 2020, 01:56:04) 
# [GCC 8.4.0]
# Embedded file name: D:\Documents\GitHub\luxpy\luxpy\utils\helpers\helpers.py
# Compiled at: 2018-06-28 06:36:21
# Size of source mod 2**32: 21052 bytes
"""
Module with helper functions 
============================

 :np2d(): Make a tuple, list or array at least a 2D numpy array.

 :np2dT(): Make a tuple, list or array at least a 2D numpy array and tranpose.

 :np3d(): Make a tuple, list or array at least a 3D numpy array.

 :np3dT(): Make a tuple, list or array at least a 3D numpy array 
           and tranpose (swap) first two axes.

 :put_args_in_db():  | Takes the args with not-None input values of a function 
                       and overwrites the values of the corresponding keys 
                       in dict db.
                     | See put_args_in_db? for more info.

 :vec_to_dict(): Convert dict to vec and vice versa.

  getdata(): Get data from csv-file or convert between pandas dataframe
             and numpy 2d-array.

 :dictkv(): Easy input of of keys and values into dict 
            (both should be iterable lists).

 :OD(): Provides a nice way to create OrderedDict "literals".

 :meshblock(): | Create a meshed block.
               | (Similar to meshgrid, but axis = 0 is retained) 
               | To enable fast blockwise calculation.

 :aplit(): Split ndarray data on (default = last) axis.

 :ajoin(): Join tuple of ndarray data on (default = last) axis.

 :broadcast_shape(): | Broadcasts shapes of data to a target_shape. 
                     | Useful for block/vector calculations when numpy fails 
                       to broadcast correctly.

 :todim(): Expand x to dimensions that are broadcast-compatable 
           with shape of another array.
           
 :write_to_excel(): Write a DataFrame to existing an Excel file into specific Sheet.

===============================================================================
"""
from luxpy import np, pd, odict, warnings
__all__ = [
 'np2d', 'np3d', 'np2dT', 'np3dT', 'put_args_in_db', 'vec_to_dict',
 'getdata', 'dictkv', 'OD', 'meshblock', 'asplit', 'ajoin',
 'broadcast_shape', 'todim', 'write_to_excel']

def np2d(data):
    """
    Make a tuple, list or numpy array at least a 2D numpy array.
    
    Args:
        :data: 
            | tuple, list, ndarray
        
    Returns:
        :returns:
            | ndarray with .ndim >= 2
    """
    if isinstance(data, np.ndarray):
        if len(data.shape) >= 2:
            return data
        else:
            return np.atleast_2d(data)
    else:
        return np.atleast_2d(np.array(data))


def np2dT(data):
    """
    Make a tuple, list or numpy array at least a 2D numpy array and transpose.
    
    Args:
        :data: 
            | tuple, list, ndarray
        
    Returns:
        :returns: 
            | ndarray with .ndim >= 2 and with transposed axes.
    """
    if isinstance(data, np.ndarray):
        if len(data.shape) >= 2:
            return data.T
        else:
            return np.atleast_2d(data).T
    else:
        return np.atleast_2d(np.array(data)).T


def np3d(data):
    """
    Make a tuple, list or numpy array at least a 3d numpy array.
    
    Args:
        :data: 
            | tuple, list, ndarray
        
    Returns:
        :returns: 
            | ndarray with .ndim >= 3
    """
    if isinstance(data, np.ndarray):
        if len(data.shape) >= 3:
            return data
        else:
            return np.expand_dims((np.atleast_2d(data)), axis=0)
    else:
        return np.expand_dims((np.atleast_2d(np.array(data))), axis=0)


def np3dT(data):
    """
    Make a tuple, list or numpy array at least a 3d numpy array and transposed 
    first 2 axes.
    
    Args:
        :data: 
            | tuple, list, ndarray
        
    Returns:
        :returns: 
            | ndarray with .ndim >= 3 and with first two axes 
              transposed (axis=3 is kept the same).
    """
    if isinstance(data, np.ndarray):
        if len(data.shape) >= 3:
            return data.transpose((1, 0, 2))
        else:
            return np.expand_dims((np.atleast_2d(data)), axis=0).transpose((1, 0, 2))
    else:
        return np.expand_dims((np.atleast_2d(np.aray(data))), axis=0).transpose((1,
                                                                                 0,
                                                                                 2))


def put_args_in_db(db, args):
    """
    Takes the args with not-None input values of a function and overwrites 
    the values of the corresponding keys in dict db.
    | (args are collected with the built-in function locals(), 
    | See example usage below)
    
    Args:
        :db: 
            | dict
    
    Returns:
        :returns: 
            | dict with the values of specific keys overwritten by the 
            |      not-None values of corresponding args of a function fcn.
    
    Example usage:
        | _db = {'c' : 'c1', 'd' : 10, 'e' : {'e1':'hello', 'e2':1000}}
        |
        | def test_put_args_in_db(a, b, db = None, c = None,d = None,e = None):
        | 
        |    args = locals().copy()  # get dict with keyword input arguments to 
        |                             # function 'test_put_args_in_db'
        |     
        |     db = put_args_in_db(db,args) # overwrite non-None args in db copy.
        |     
        |    if db is not None: # unpack db for further use
        |        c,d,e = [db[x] for x in sorted(db.keys())]
        |     
        |     print(' a : {}'.format(a))
        |     print(' b : {}'.format(b))
        |     print(' db: {}'.format(db))
        |     print(' c : {}'.format(c))
        |     print(' d : {}'.format(d))
        |     print(' e : {}'.format(e))
        |     print('_db: {}'.format(_db))
 
    """
    if db is not None:
        dbc = db.copy()
        dbc_keys = dbc.keys()
        for argkey in args.keys():
            if args[argkey] is not None and argkey in dbc_keys:
                dbc[argkey] = args[argkey]

        return dbc
    else:
        return db


def vec_to_dict(vec=None, dic={}, vsize=None, keys=None):
    """
    Convert dict to vec and vice versa.
    
    Args:
        :vec: 
            | list or vector array, optional
        :dic: 
            | dict, optional
        :vsize:
            | list or vector array with size of values of dict, optional
        :keys:
            | list or vector array with keys in dict (must be provided).
        
    Returns:
        :returns:
            | x, vsize
            |   x is an array, if vec is None
            |   x is a dict, if vec is not None
    """
    if vec is not None:
        n = 0
        for i, v in enumerate(keys):
            dic[v] = vec[(n + np.arange(vsize[i]))]
            n += dic[v].shape[0]

        return (dic, vsize)
    else:
        vec = []
        vsize = []
        for i, v in enumerate(keys):
            vec = np.hstack((vec, dic[v]))
            vsize.append(dic[v].shape[0])

        return (
         vec, vsize)


def getdata(data, kind='np', columns=None, header=None, sep=',', datatype='S', verbosity=True):
    """
    Get data from csv-file 
    or convert between pandas dataframe and numpy 2d-array.
    
    Args:
        :data: 
            | - str with path to file containing data
            | - ndarray with data
            | - pandas.dataframe with data
        :kind: 
            | str ['np','df'], optional 
            | Determines type(:returns:), np: ndarray, df: pandas.dataframe
        :columns:
            | None or list[str] of column names for dataframe, optional
        :header:
            | None, optional
            |   - None: no header in file
            |   - 'infer': infer headers from file
        :sep:
            | ',' or '  ' or other char, optional
            | Column separator in data file
        :datatype':
            | 'S',optional 
            | Specifies a type of data. 
            | Is used when creating column headers (:column: is None).
            |   -'S': light source spectrum
            |   -'R': reflectance spectrum
            |   or other.      
        :verbosity:
            | True, False, optional
            | Print warning when inferring headers from file.
    
    Returns:
        :returns:
            | data as ndarray or pandas.dataframe
      
    """
    if isinstance(data, str):
        datafile = data
        data = pd.read_csv(data, names=None, index_col=None, header=header, sep=sep)
        if header == 'infer':
            if verbosity == True:
                warnings.warn('getdata(): Infering HEADERS from data file: {}!'.format(datafile))
                columns = data.columns
        else:
            if columns is None:
                data.columns = ['{}{}'.format(datatype, x) for x in range(len(data.columns))]
            if columns is not None:
                data.columns = columns
    if isinstance(data, np.ndarray) & (kind == 'df'):
        if columns is None:
            columns = ['{}{}'.format(datatype, x) for x in range(data.shape[1])]
        data = pd.DataFrame(data, columns=columns)
    else:
        if isinstance(data, pd.DataFrame) & (kind == 'np'):
            data = data.values
    return data


def dictkv(keys=None, values=None, ordered=True):
    """
    Easy input of of keys and values into dict.
    
    Args:
        :keys: 
            | iterable list[str,...] of keys
        :values:
            | iterable list[...,..., ] of values
        :ordered:
            | True, False, optional
            | True: creates an ordered dict using 'collections.OrderedDict()'
    Returns:
        :returns:
            | (ordered) dict
    """
    if ordered is True:
        return odict(zip(keys, values))
    else:
        return dict(zip(keys, values))


class OD(object):
    __doc__ = '\n    This class provides a nice way to create OrderedDict "literals".\n    \n    '

    def __getitem__(self, slices):
        if not isinstance(slices, tuple):
            slices = (
             slices,)
        return odict((slice.start, slice.stop) for slice in slices)


OD = OD()

def meshblock(x, y):
    """
    Create a meshed block from x and y.
    
    | (Similar to meshgrid, but axis = 0 is retained).
    | To enable fast blockwise calculation.
    
    Args: 
        :x: 
            | ndarray with ndim == 2
        :y: 
            | ndarray with ndim == 2
        
    Returns:
        :X,Y: 
            | 2 ndarrays with ndim == 3 
            |   X.shape = (x.shape[0],y.shape[0],x.shape[1])
            |   Y.shape = (x.shape[0],y.shape[0],y.shape[1])
    """
    Y = np.transpose((np.repeat(y, (x.shape[0]), axis=0).reshape((y.shape[0], x.shape[0], y.shape[1]))), axes=[1, 0, 2])
    X = np.repeat(x, (y.shape[0]), axis=0).reshape((x.shape[0], y.shape[0], x.shape[1]))
    return (X, Y)


def asplit(data):
    """
    Split data on last axis
    
    Args:
        :data: 
            | ndarray
        
    Returns:
        :returns: 
            | ndarray, ndarray, ... 
            |   (number of returns is equal data.shape[-1])
    """
    return [data[(..., x)] for x in range(data.shape[(-1)])]


def ajoin(data):
    """
    Join data on last axis.
    
    Args:
        :data:
            | tuple (ndarray, ndarray, ...)
        
    Returns:
        :returns:
            | ndarray (shape[-1] is equal to tuple length)
    """
    if data[0].ndim == 2:
        return np.transpose(np.concatenate(data, axis=0).reshape(np.hstack((len(data), data[0].shape))), (1,
                                                                                                          2,
                                                                                                          0))
    else:
        if data[0].ndim == 1:
            return np.concatenate(data, axis=0).reshape(np.hstack((len(data), data[0].shape))).T
        return np.hstack(data)[0]


def broadcast_shape(data, target_shape=None, expand_2d_to_3d=None, axis0_repeats=None, axis1_repeats=None):
    """
    Broadcasts shapes of data to a target_shape.
    
    | Useful for block/vector calc. when numpy fails to broadcast correctly.
    
    Args:
        :data: 
            | ndarray 
        :target_shape: 
            | None or tuple with requested shape, optional
            |   - None: returns unchanged :data:
        :expand_2d_to_3d:
            | None (do nothing) or ..., optional 
            | If ndim == 2, expand from 2 to 3 dimensions
        :axis0_repeats:
            | None or number of times to repeat axis=0, optional
            |   - None: keep axis=0 same size
        :axis1_repeats:
            | None or number of times to repeat axis=1, optional
            |   - None: keep axis=1 same size

    Returns:
        :returns: 
            | reshaped ndarray
    """
    data = np2d(data)
    if (expand_2d_to_3d is not None) & (len(data.shape) == 2):
        data = np.expand_dims(data, axis=expand_2d_to_3d)
    if target_shape is not None:
        dshape = data.shape
        if dshape != target_shape:
            axis_of_v3 = len(target_shape) - 1
            if dshape[0] != target_shape[0]:
                if axis0_repeats is None:
                    axis0_repeats = target_shape[0] - dshape[0] + 1
                data = np.repeat(data, axis0_repeats, axis=0)
            if (len(target_shape) > 2) & (len(data.shape) == 2):
                data = np.expand_dims(data, axis=(axis_of_v3 - 1))
                dshape = data.shape
                if dshape[1] != target_shape[1]:
                    if axis1_repeats is None:
                        axis1_repeats = target_shape[1] - dshape[1] + 1
                    data = np.repeat(data, axis1_repeats, axis=1)
        for i in range(2):
            if (data.shape[i] > 1) & (data.shape[i] != target_shape[i]):
                raise Exception('broadcast_shape(): Cannot match dim of data with target: data.shape[i]>1  & ...  != target.shape[i]')

    return data


def todim(x, tshape, add_axis=1, equal_shape=False):
    """
    Expand x to dims that are broadcast-compatable with shape of another array.
    
    Args:
        :x: 
            | ndarray
        :tshape: 
            | tuple with target shape
        :add_axis:
            | 1, optional
            | Determines where in x.shape an axis should be added
        :equal_shape:
            | False or True, optional
            | True: expand :x: to identical dimensions (speficied by :tshape:)
            
    Returns:
        :returns:
            | ndarray broadcast-compatable with tshape.
    """
    if x is None:
        return np.broadcast_arrays(x, np.ones(tshape))[0]
    else:
        x = np2d(x)
        sx = x.shape
        lsx = len(sx)
        ltshape = len(tshape)
        if sx == tshape:
            pass
        else:
            if (lsx == 1) | (sx == (1, tshape[(-1)])) | (sx == (tshape[(-1)], 1)):
                if sx == (tshape[(-1)], 1):
                    x = x.T
                if lsx != ltshape:
                    x = np.expand_dims(x, 0)
            else:
                if lsx == 2:
                    if ltshape == 3:
                        sd = np.setdiff1d(tshape, sx, assume_unique=True)
                        if len(sd) == 0:
                            ax = add_axis
                        else:
                            ax = np.where(tshape == sd)[0][0]
                        x = np.expand_dims(x, ax)
                    else:
                        raise Exception('todim(x,tshape): dimensions do not match for 2d arrays.')
                else:
                    raise Exception('todim(x,tshape): no matching dimensions between 3d x and tshape.')
        if equal_shape == False:
            return x
        return np.ones(tshape) * x


def write_to_excel(filename, df, sheet_name='Sheet1', startrow=None, truncate_sheet=False, **to_excel_kwargs):
    """
    Writes a DataFrame to an existing Excel file into a specified sheet.
    | If [filename] doesn't exist, then this function will create it.

    Args:
      :filename: 
          | File path or existing ExcelWriter
          | (Example: '/path/to/file.xlsx')
      :df: 
          | dataframe to save to workbook
      :sheet_name: 
          | Name of sheet which will contain DataFrame.
          | (default: 'Sheet1')
      :startrow: 
          | upper left cell row to dump data frame.
          | Per default (startrow=None) calculate the last row
          | in the existing DF and write to the next row...
      :truncate_sheet: 
          | truncate (remove and recreate) [sheet_name]
          | before writing DataFrame to Excel file
      :to_excel_kwargs: 
          | arguments which will be passed to `DataFrame.to_excel()`
          | [can be dictionary]

    Returns: None
    
    Notes:
        Copied from https://stackoverflow.com/questions/20219254/how-to-write-to-an-existing-excel-file-without-overwriting-data-using-pandas
    """
    from openpyxl import load_workbook
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')
    writer = pd.ExcelWriter(filename, engine='openpyxl')
    try:
        writer.book = load_workbook(filename)
        if startrow is None:
            if sheet_name in writer.book.sheetnames:
                startrow = writer.book[sheet_name].max_row
        if truncate_sheet:
            if sheet_name in writer.book.sheetnames:
                idx = writer.book.sheetnames.index(sheet_name)
                writer.book.remove(writer.book.worksheets[idx])
                writer.book.create_sheet(sheet_name, idx)
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        pass

    if startrow is None:
        startrow = 0
    (df.to_excel)(writer, sheet_name, startrow=startrow, **to_excel_kwargs)
    writer.save()