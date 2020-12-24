# uncompyle6 version 3.7.4
# Python bytecode 2.7 (62211)
# Decompiled from: Python 3.6.9 (default, Apr 18 2020, 01:56:04) 
# [GCC 8.4.0]
# Embedded file name: /mnt/d/Sandbox/huru-server/.venv/lib/python2.7/site-packages/data_importer/readers/xlsx_reader.py
# Compiled at: 2020-04-17 10:46:24
from __future__ import unicode_literals
from openpyxl import load_workbook

class XLSXReader(object):

    def __init__(self, instance, data_only=True, sheet_index=None):
        self.workbook = load_workbook(instance.source, data_only=data_only)
        if instance.Meta.sheet_name:
            self.worksheet = self.workbook.get_sheet_by_name(instance.Meta.sheet_name)
        else:
            if sheet_index is None:
                sheet_index = instance.Meta.sheet_index or 0
            self.worksheet = self.workbook.worksheets[sheet_index]
        return

    def read(self):
        for line, row in enumerate(self.worksheet.iter_rows()):
            values = [ cell.value for cell in row ]
            yield values