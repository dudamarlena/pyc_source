# uncompyle6 version 3.7.4
# Python bytecode 3.6 (3379)
# Decompiled from: Python 3.6.9 (default, Apr 18 2020, 01:56:04) 
# [GCC 8.4.0]
# Embedded file name: /home/turicas/software/pyenv/versions/rows/lib/python3.6/site-packages/rows/plugins/xlsx.py
# Compiled at: 2019-02-13 03:47:25
# Size of source mod 2**32: 6038 bytes
from __future__ import unicode_literals
from decimal import Decimal
from io import BytesIO
from numbers import Number
from openpyxl import Workbook, load_workbook
from openpyxl.cell.read_only import EmptyCell
from openpyxl.utils import get_column_letter
from rows import fields
from rows.plugins.utils import create_table, get_filename_and_fobj, prepare_to_export

def _cell_to_python(cell):
    """Convert a PyOpenXL's `Cell` object to the corresponding Python object."""
    data_type, value = cell.data_type, cell.value
    if type(cell) is EmptyCell:
        return
    if data_type == 'f':
        if value == '=TRUE()':
            return True
    if data_type == 'f':
        if value == '=FALSE()':
            return False
    if cell.number_format.lower() == 'yyyy-mm-dd':
        return str(value).split(' 00:00:00')[0]
    if cell.number_format.lower() == 'yyyy-mm-dd hh:mm:ss':
        return str(value).split('.')[0]
    if cell.number_format.endswith('%') and isinstance(value, Number):
        value = Decimal(str(value))
        return '{:%}'.format(value)
    else:
        if value is None:
            return ''
        return value


def sheet_cell(sheet, row, col):
    return sheet[(get_column_letter(col) + str(row))]


def import_from_xlsx(filename_or_fobj, sheet_name=None, sheet_index=0, start_row=None, start_column=None, end_row=None, end_column=None, *args, **kwargs):
    """Return a rows.Table created from imported XLSX file."""
    workbook = load_workbook(filename_or_fobj, read_only=True)
    if sheet_name is None:
        sheet_name = workbook.sheetnames[sheet_index]
    sheet = workbook[sheet_name]
    min_row, min_column = sheet.min_row - 1, sheet.min_column - 1
    max_row, max_column = sheet.max_row - 1, sheet.max_column - 1
    start_row = start_row if start_row is not None else min_row
    end_row = end_row if end_row is not None else max_row
    start_column = start_column if start_column is not None else min_column
    end_column = end_column if end_column is not None else max_column
    table_rows = []
    is_empty = lambda row: all(cell is None for cell in row)
    for row_index in range(start_row + 1, end_row + 2):
        row = [_cell_to_python(sheet_cell(sheet, row_index, col_index)) for col_index in range(start_column + 1, end_column + 2)]
        if not is_empty(row):
            table_rows.append(row)

    filename, _ = get_filename_and_fobj(filename_or_fobj, dont_open=True)
    metadata = {'imported_from':'xlsx',  'filename':filename,  'sheet_name':sheet_name}
    return create_table(table_rows, *args, meta=metadata, **kwargs)


FORMATTING_STYLES = {fields.DateField: 'YYYY-MM-DD', 
 fields.DatetimeField: 'YYYY-MM-DD HH:MM:SS', 
 fields.PercentField: '0.00%'}

def _python_to_cell(field_types):

    def convert_value(field_type, value):
        number_format = FORMATTING_STYLES.get(field_type, None)
        if field_type not in (
         fields.BoolField,
         fields.DateField,
         fields.DatetimeField,
         fields.DecimalField,
         fields.FloatField,
         fields.IntegerField,
         fields.PercentField,
         fields.TextField):
            value = field_type.serialize(value)
        return (value, number_format)

    def convert_row(row):
        return [convert_value(field_type, value) for field_type, value in zip(field_types, row)]

    return convert_row


def export_to_xlsx(table, filename_or_fobj=None, sheet_name='Sheet1', *args, **kwargs):
    """Export the rows.Table to XLSX file and return the saved file."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    prepared_table = prepare_to_export(table, *args, **kwargs)
    field_names = next(prepared_table)
    for col_index, field_name in enumerate(field_names):
        cell = sheet.cell(row=1, column=(col_index + 1))
        cell.value = field_name

    _convert_row = _python_to_cell(list(map(table.fields.get, field_names)))
    for row_index, row in enumerate(prepared_table, start=1):
        for col_index, (value, number_format) in enumerate(_convert_row(row)):
            cell = sheet.cell(row=(row_index + 1), column=(col_index + 1))
            cell.value = value
            if number_format is not None:
                cell.number_format = number_format

    if filename_or_fobj is not None:
        _, fobj = get_filename_and_fobj(filename_or_fobj, mode='wb')
        workbook.save(fobj)
        fobj.flush()
        return fobj
    else:
        fobj = BytesIO()
        workbook.save(fobj)
        fobj.seek(0)
        result = fobj.read()
        fobj.close()
        return result