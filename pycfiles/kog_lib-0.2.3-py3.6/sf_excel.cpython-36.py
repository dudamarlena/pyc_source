# uncompyle6 version 3.7.4
# Python bytecode 3.6 (3379)
# Decompiled from: Python 3.6.9 (default, Apr 18 2020, 01:56:04) 
# [GCC 8.4.0]
# Embedded file name: build\bdist.win-amd64\egg\kog_lib\excel\sf_excel.py
# Compiled at: 2020-01-07 04:59:41
# Size of source mod 2**32: 10714 bytes
import os
from pprint import pprint
import xlrd, xlwt
from xlutils.copy import copy
from openpyxl import load_workbook, Workbook

class Excel:

    def __init__(self, file=None, index=0):
        self.rowNum = 0
        self.colNum = 0
        self.data = list()
        if file:
            if not os.path.isfile(file):
                raise FileNotFoundError('文件{}不存在'.format(file))
            else:
                self.load(file, index)

    @staticmethod
    def help():
        print('\n        ----------------------Excel类 使用方法--------------------------------------\n        import kog_lib\n\n        实例化：\n            my_excel = kog_lib.Excel(excel_path)\n            或，\n            my_excel = kog_lib.Excel()\n            my_excel.load(excel_path)\n\n        数据读取：\n            my_excel.data = [\n                [第一行],\n                [第二行],\n                ...\n                [最后一行]\n            ]\n\n        写入数据：\n            my_excel.make(data, filename=\'name1.xls\', sheetname=\'sheet1\')\n                data是一个二位的列表，如上述“数据读取”部分的“my_excel.data”\n\n        打印内容：\n            my_excel.print(part,n)\n                part: "all" 、"head" 、"tail"：全部打印、 从头打印n个、 从尾部打印n个\n                n: 打印个数\n        ----------------------------------------------------------------------------\n            ')

    def load(self, file: str, index=0):
        """'
        file:   str, 表示要加载的文件
        index： (int,str), 表示要加载的sheet
        """
        if not os.path.isfile(file):
            raise FileNotFoundError('文件不存在或不是文件')
        else:
            try:
                end = file.rsplit('.', 1)[(-1)]
            except Exception as e:
                raise TypeError(f"文件格式不准确, {e}")

            if end == 'xls':
                self.xls_load_list(file, index)
            else:
                if end in ('xlsx', 'xlsm', 'xltx', 'xltm'):
                    self.xlsx_load_list(file, index)
                else:
                    raise TypeError('文件格式不正确')
        print(f"文件{file}加载完毕～")

    def xlsx_load_list(self, file: str, index: int=0):
        """

        加载xlsx格式的文件
        :param file:
        :param index:
        :return:
        """
        wb = load_workbook(filename=file, data_only=True)
        sheet_names = wb.get_sheet_names()
        if index < len(sheet_names):
            ws = wb.get_sheet_by_name(sheet_names[index])
        else:
            raise IndexError(f"输入的页码{index}不存在")
        rows = ws.rows
        columns = ws.columns
        _list = []
        for i, row in enumerate(rows):
            rlist = []
            for j, column in enumerate(row):
                rlist.append(column.value)

            _list.append(rlist)

        self.data = _list
        self.rowNum = ws.max_row
        self.colNum = ws.max_column

    def xls_load_list(self, file: str, index: int=0):
        workbook = xlrd.open_workbook(file)
        sheets = workbook.sheets()
        if isinstance(index, int):
            if index < len(sheets):
                sheet = workbook.sheets()[index]
            else:
                raise IndexError(f"输入的页码{index}不存在")
        else:
            if isinstance(index, str):
                sheet = workbook.sheet_by_name(index)
            else:
                raise TypeError(f"{index}参数错误～")
        print(f"正在加载文件{file}...")
        self.rowNum = sheet.nrows
        self.colNum = sheet.ncols
        _arr = []
        for i in range(self.rowNum):
            rowlist = []
            for j in range(self.colNum):
                rowlist.append(sheet.cell_value(i, j))

            _arr.append(rowlist)

        self.data = _arr

    def data_validater(self, data):
        """
        判断要写入的data是否满足条件，可以写入
        :param data: 要判断的数据
        :return: Boolean
        """
        if not isinstance(data, (list, tuple)):
            return False
        else:
            for item in data:
                if not isinstance(item, (list, tuple)):
                    return False

            return True

    def __write_sheet(self, my_excel, active_sheet, filename, data):
        """
        写入数据到excel的方法
        :param my_excel: xlwt或openpyxl的实例
        :param active_sheet: 要写入的工作簿的实例
        :param filename: 要写入的文件名
        :param data: 要判断的数据
        :return: void
        """
        end = filename.rsplit('.', 1)[(-1)]
        if end == 'xls':
            for i in range(len(data)):
                for j in range(len(data[i])):
                    active_sheet.write(i, j, (f"{data[i][j]}"))

            my_excel.save(filename)
        if end in ('xlsx', 'xlsm', 'xltx', 'xltm'):
            for l1 in data:
                active_sheet.append(l1)

            my_excel.save(filename=filename)

    def __get_sheetname(self, my_excel, mode):
        """
        获取工作簿名的列表
        :param my_excel: xlwt或openpyxl的实例
        :param data: 原文件是xls，还是xlsx
        :return: list
        """
        namelist = []
        if mode == 'xls':
            for sheet in my_excel.sheets():
                namelist.append(sheet.name)

            return namelist
        if mode == 'xlsx':
            return my_excel.get_sheet_names()

    def xls_write_new_file(self, data, filename, sheetname):
        """
        写入一个新的文件
        :param data: 要写入的数据
        :param filename: 要写入的文件名
        :param sheetname: 要写入的工作簿名
        :return: void
        """
        my_excel = xlwt.Workbook()
        my_sheet = my_excel.add_sheet(sheetname=sheetname, cell_overwrite_ok=True)
        self._Excel__write_sheet(my_excel, my_sheet, filename, data)

    def xls_add_sheet_2_file(self, data, filename, sheetname):
        """
        写入一个写入现有excel文件中，追加一个新的工作簿
        :param data: 要写入的数据
        :param filename: 要写入的文件名
        :param sheetname: 要写入的工作簿名
        :return: void
        """
        my_excel = xlrd.open_workbook(filename, formatting_info=True)
        if sheetname in self._Excel__get_sheetname(my_excel, 'xls'):
            raise ValueError(f"工作簿名({sheetname})已存在")
        my_excel = copy(wb=my_excel)
        my_sheet = my_excel.add_sheet(sheetname=sheetname, cell_overwrite_ok=True)
        self._Excel__write_sheet(my_excel, my_sheet, filename, data)

    def xlsx_write_new_file(self, data, filename, sheetname):
        """
        写入一个写入现有excel文件中，追加一个新的工作簿
        :param data: 要写入的数据
        :param filename: 要写入的文件名
        :param sheetname: 要写入的工作簿名
        :return: void
        """
        my_excel = Workbook()
        my_sheet = my_excel.active
        my_sheet.title = sheetname
        self._Excel__write_sheet(my_excel, my_sheet, filename, data)

    def xlsx_add_sheet_2_file(self, data, filename, sheetname):
        """
        写入一个写入现有excel文件中，追加一个新的工作簿
        :param data: 要写入的数据
        :param filename: 要写入的文件名
        :param sheetname: 要写入的工作簿名
        :return: void
        """
        my_excel = load_workbook(filename)
        my_sheet = my_excel.create_sheet()
        if sheetname in self._Excel__get_sheetname(my_excel, 'xlsx'):
            raise ValueError(f"工作簿名({sheetname})已存在")
        my_sheet.title = sheetname
        self._Excel__write_sheet(my_excel, my_sheet, filename, data)

    def make(self, data, filename='name1.xls', sheetname='sheet1'):
        """生成一个excel文件"""
        end = filename.rsplit('.', 1)[(-1)]
        if end not in ('xls', 'xlsx', 'xlsm', 'xltx', 'xltm'):
            raise TypeError('文件格式不正确, 只支持xls, xlsx, xlsm, xltx, xltm格式')
        elif not self.data_validater(data):
            raise TypeError(f"data({type(data)})不是列表，或者data列表的元素不是列表")
        else:
            if end == 'xls':
                if len(data) > 65536:
                    raise ValueError(f"总行数({len(data)})超出上限,会造成数据丢失，请分工作簿写入")
                else:
                    if not os.path.exists(filename):
                        self.xls_write_new_file(data, filename, sheetname)
                    else:
                        self.xls_add_sheet_2_file(data, filename, sheetname)
            else:
                if end in ('xlsx', 'xlsm', 'xltx', 'xltm'):
                    if len(data) > 1048576:
                        raise ValueError(f"总行数({len(data)})超出上限,会造成数据丢失，请分工作簿写入")
                    else:
                        if not os.path.exists(filename):
                            self.xlsx_write_new_file(data, filename, sheetname)
                        else:
                            self.xlsx_add_sheet_2_file(data, filename, sheetname)
                else:
                    raise TypeError('文件格式不正确, 只支持xls, xlsx, xlsm, xltx, xltm格式')

    def __str__(self):
        return str(vars(self))

    def print(self, part='all', n=100):
        """
        打印出来
        :return:
        """
        print('行：', self.rowNum, '列：', self.colNum)
        if part == 'all':
            for i in range(self.rowNum):
                for j in range(self.colNum):
                    print((self.data[i][j]), '\t\t', end='')

                print()

        else:
            if part == 'head':
                n = 0 if n - 1 <= 0 else self.rowNum if n - 1 >= self.rowNum else n
                for i in range(n):
                    for j in range(self.colNum):
                        print((self.data[i][j]), '\t\t', end='')

                    print()

            else:
                if part == 'tail':
                    n = 0 if n - 1 <= 0 else self.rowNum if n - 1 >= self.rowNum else n
                    for i in range(n - 1, self.rowNum):
                        for j in range(self.colNum):
                            print((self.data[i][j]), '\t\t', end='')

                        print()

                else:
                    raise ValueError(f"part参数{part}错误")

    def print_list(self):
        pprint(self.data)