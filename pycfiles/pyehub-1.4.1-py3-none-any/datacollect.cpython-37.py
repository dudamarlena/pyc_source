# uncompyle6 version 3.6.7
# Python bytecode 3.7 (3394)
# Decompiled from: Python 3.8.2 (tags/v3.8.2:7b3ab59, Feb 25 2020, 23:03:10) [MSC v.1916 64 bit (AMD64)]
# Embedded file name: /Users/juancomish/miniconda3/lib/python3.7/site-packages/pyehub/datacollect.py
# Compiled at: 2019-07-03 19:21:52
# Size of source mod 2**32: 648 bytes
__doc__ = '\nA script for printing out the contents of an excel spreadsheet.\n'
import openpyxl
from openpyxl import load_workbook

def main():
    """Main function."""
    book = openpyxl.load_workbook('Individual_Hubs/LI_Batresuls_11.xlsx')
    print(book.get_sheet_names())
    sheet = book.get_sheet_by_name('Other')
    num_row = sheet.max_row
    num_column = sheet.max_column
    for i in range(1, num_row + 1):
        for j in range(1, num_column + 1):
            print(sheet.cell(row=i, column=j).value)


if __name__ == '__main__':
    main()