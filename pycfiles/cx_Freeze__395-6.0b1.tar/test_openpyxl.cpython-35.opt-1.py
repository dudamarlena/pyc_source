# uncompyle6 version 3.6.7
# Python bytecode 3.5 (3351)
# Decompiled from: Python 3.8.2 (tags/v3.8.2:7b3ab59, Feb 25 2020, 23:03:10) [MSC v.1916 64 bit (AMD64)]
# Embedded file name: \.\cx_Freeze\samples\openpyxl\test_openpyxl.py
# Compiled at: 2019-08-29 22:24:38
# Size of source mod 2**32: 574 bytes
from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws['A1'] = 42
ws.append([1, 2, 3])
import datetime
ws['A2'] = datetime.datetime.now()
fileName = 'sample.xlsx'
wb.save(fileName)
print('Wrote file', fileName)