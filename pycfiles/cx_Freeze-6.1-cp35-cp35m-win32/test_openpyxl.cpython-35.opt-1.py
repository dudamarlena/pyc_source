# uncompyle6 version 3.7.4
# Python bytecode 3.5 (3351)
# Decompiled from: Python 3.6.9 (default, Apr 18 2020, 01:56:04) 
# [GCC 8.4.0]
# Embedded file name: \.\cx_Freeze\samples\openpyxl\test_openpyxl.py
# Compiled at: 2019-08-29 22:24:38
# Size of source mod 2**32: 574 bytes
from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws['A1'] = 42
ws.append([1, 2, 3])
import datetime
ws['A2'] = datetime.datetime.now()
fileName = 'sample.xlsx'
wb.save(fileName)
print('Wrote file', fileName)