# uncompyle6 version 3.7.4
# Python bytecode 3.6 (3379)
# Decompiled from: Python 3.6.9 (default, Apr 18 2020, 01:56:04) 
# [GCC 8.4.0]
# Embedded file name: build/bdist.linux-x86_64/egg/basico/services/srv_bnr.py
# Compiled at: 2019-04-01 16:37:59
# Size of source mod 2**32: 13820 bytes
"""
# File: srv_bnr.py
# Author: Tomás Vírseda
# License: GPL v3
# Description: Backup and Restore service
"""
import os
from os import linesep as NEWLINE
from os.path import sep as SEP
import csv, json, glob, shutil
from datetime import datetime
from tempfile import TemporaryDirectory
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, Border, Side
from basico.core.mod_srv import Service
from basico.core.mod_env import LPATH, APP
HEADER_FIELDS = [
 'id', 'version', 'title', 'componentkey', 'componenttxt', 'category', 'priority', 'type', 'releasedon', 'collections']
COL = {1:'A', 
 2:'B', 
 3:'C', 
 4:'D', 
 5:'E', 
 6:'F', 
 7:'G', 
 8:'H', 
 9:'I', 
 10:'J'}
HEADER = {'id':'SAP Note Id', 
 'version':'Version', 
 'title':'Title', 
 'componentkey':'Component', 
 'componenttxt':'Component Description', 
 'category':'Category', 
 'priority':'Priority', 
 'type':'Type', 
 'releasedon':'Released On', 
 'collections':'Collections'}

class BackupRestoreMan(Service):

    def initialize(self):
        self.get_services()

    def get_services(self):
        self.srvstg = self.get_service('Settings')
        self.srvdtb = self.get_service('DB')
        self.srvgui = self.get_service('GUI')
        self.srvuif = self.get_service('UIF')
        self.srvsap = self.get_service('SAP')
        self.srvicm = self.get_service('IM')
        self.srvutl = self.get_service('Utils')
        self.srvant = self.get_service('Annotation')
        self.srvclt = self.get_service('Collections')

    def backup(self, bck_file, backup_annotations):
        with TemporaryDirectory(dir=(LPATH['TMP'])) as (TEMPORARY_DIRECTORY):
            SOURCE_DIRECTORY = LPATH['DB']
            self.log.debug('Copying database structure to a temp directory:')
            self.log.debug('From: %s', SOURCE_DIRECTORY)
            self.log.debug('  To: %s', TEMPORARY_DIRECTORY)
            shutil.rmtree(TEMPORARY_DIRECTORY)
            shutil.copytree(SOURCE_DIRECTORY, TEMPORARY_DIRECTORY)
            self.log.debug('Contents copied successfully')
            if not backup_annotations:
                self.log.debug('Deleting annotations directory in temporary foleder:')
                TEMPORARY_ANNOTATION_DIRECTORY = TEMPORARY_DIRECTORY + SEP + 'annotations'
                self.log.debug(TEMPORARY_ANNOTATION_DIRECTORY)
                shutil.rmtree(TEMPORARY_ANNOTATION_DIRECTORY)
                self.log.debug('Annotations deleted successfully')
            bckname = self.srvutl.zip(bck_file, TEMPORARY_DIRECTORY)
            self.log.info('Database successfully backed up in: %s' % bckname)
            return bckname

    def test(self, bckfile=None):
        self.log.info('Testing backup: %s', bckfile)
        try:
            with TemporaryDirectory(dir=(LPATH['TMP'])) as (tmpdir):
                self.log.debug('Temporary directory for testing: %s', tmpdir)
                self.srvutl.unzip(bckfile, tmpdir)
                fsapnotes = tmpdir + SEP + 'sapnotes' + SEP + 'sapnotes.json'
                with open(fsapnotes, 'r') as (fp):
                    sapnotes = json.load(fp)
                ANNOTATIONS_DIR = tmpdir + SEP + 'annotations' + SEP
                annotations = glob.glob(ANNOTATIONS_DIR + '*.json')
                fcols = tmpdir + SEP + 'collections' + SEP + 'collections.json'
                with open(fcols, 'r') as (fp):
                    cols = json.load(fp)
                self.log.info('Found %d SAP Notes, %d collections and %d annotations in this backup', len(sapnotes), len(cols), len(annotations))
                return (len(sapnotes), len(annotations), len(cols))
        except Exception as error:
            self.log.error(error)
            self.log.error(self.get_traceback())
            return

    def get_all_backups(self):
        backups = glob.glob(LPATH['BACKUP'] + '*.zip')
        backups.sort(reverse=True)
        return backups

    def reorg(self):
        self.log.info('Backup reorganization')
        backups = self.get_all_backups()
        if len(backups) > 9:
            n = 0
            for backup in backups[9:]:
                os.unlink(backup)
                self.log.info('\tBackup %s deleted' % os.path.basename(backup))
                n += 1

            self.log.info('\tDeleted %d backups' % n)

    def restore_from_backup(self, bckfile, overwrite=False):
        self.log.info('Restoring backup: %s (Overwrite mode is: %s)', bckfile, overwrite)
        try:
            with TemporaryDirectory(dir=(LPATH['TMP'])) as (tmpdir):
                self.log.debug('Using temporary dir for import: %s', tmpdir)
                self.srvutl.unzip(bckfile, tmpdir)
                fimpnotes = tmpdir + SEP + 'sapnotes' + SEP + 'sapnotes.json'
                with open(fimpnotes, 'r') as (fp):
                    impnotes = json.load(fp)
                    self.log.info('Found %d SAP Notes in backup' % len(impnotes))
                    imported = self.srvdtb.add_list(impnotes, overwrite)
                    self.log.info('Imported %d SAP Notes', imported)
                fimpcols = tmpdir + SEP + 'collections' + SEP + 'collections.json'
                with open(fimpcols, 'r') as (fp):
                    impcols = json.load(fp)
                    for cid in impcols:
                        self.srvclt.create((impcols[cid]), cid, batch=True)

                    self.srvclt.save()
                    self.log.info('Imported %d collections from backup' % len(impcols))
                fimpannot = tmpdir + SEP + 'annotations'
                if os.path.exists(fimpannot):
                    source = fimpannot + SEP + '*'
                    annotations_files = glob.glob(source)
                    target = LPATH['ANNOTATIONS']
                    for filename in annotations_files:
                        shutil.copy(filename, target)

                    self.log.info('Imported %d annotations from backup', int(len(annotations_files) / 2))
        except Exception as error:
            self.log.error(error)
            self.log.debug(self.get_traceback())

    def restore_from_cache(self, *args):
        valid = 0
        db = self.get_service('DB')
        sap = self.get_service('SAP')
        xml_files = glob.glob(LPATH['CACHE_XML'] + '*.xml')
        for xml in xml_files:
            sid = xml[xml.rfind(SEP) + 1:xml.rfind('.')]
            content = open(xml, 'r').read()
            sapnote = sap.analyze_sapnote(sid, content)
            try:
                sapnote['id']
                db.add(sapnote)
                valid += 1
            except:
                self.log.error('SAP Note %s is not valid. Skipping.' % sid)

        self.log.info('Restored %d of %d available files' % (valid, len(xml_files)))
        db.save_notes()

    def export_to_text_csv(self, bag, export_path=None):
        if len(bag) == 0:
            return
        try:
            writer = csv.writer((open(export_path, 'w')), delimiter=';', quoting=(csv.QUOTE_ALL))
            csvrow = []
            for field in HEADER_FIELDS:
                csvrow.append(HEADER[field])

            writer.writerow(csvrow)
            for sid in bag:
                csvrow = []
                metadata = self.srvdtb.get_sapnote_metadata(sid)
                for field in HEADER_FIELDS:
                    if field == 'collections':
                        cols = ', '.join([self.srvclt.get_name_by_cid(col) for col in metadata[field]])
                        csvrow.append(cols)
                    else:
                        if field == 'releasedon':
                            excel_date = self.srvutl.get_excel_date(metadata[field])
                            csvrow.append(excel_date)
                        else:
                            csvrow.append(str(metadata[field]))

                writer.writerow(csvrow)

            return True
        except Exception as error:
            self.log.error(error)
            return False

    def export_to_basico(self, bag, export_path=None):
        if len(bag) == 0:
            return
        try:
            bag_notes = {}
            sapnotes = self.srvdtb.get_notes()
            for sid in bag:
                sid = self.srvdtb.normalize_sid(sid)
                bag_notes[sid] = sapnotes[sid]

            bag_cols = {}
            collections = self.srvclt.get_all()
            for sid in bag_notes:
                cols = bag_notes[sid]['collections']
                for cid in cols:
                    bag_cols[cid] = collections[cid]

            with TemporaryDirectory(dir=(LPATH['TMP'])) as (tmpdir):
                self.log.debug('Using temporary dir for export: %s', tmpdir)
                os.makedirs(tmpdir + SEP + 'sapnotes')
                os.makedirs(tmpdir + SEP + 'collections')
                os.makedirs(tmpdir + SEP + 'annotations')
                os.makedirs(tmpdir + SEP + 'resources')
                bagsnfile = tmpdir + SEP + 'sapnotes' + SEP + 'sapnotes.json'
                with open(bagsnfile, 'w') as (fsn):
                    json.dump(bag_notes, fsn)
                    self.log.debug('Saved %d notes to %s' % (len(bag_notes), bagsnfile))
                bagcolsfile = tmpdir + SEP + 'collections' + SEP + 'collections.json'
                with open(bagcolsfile, 'w') as (fcols):
                    json.dump(bag_cols, fcols)
                    self.log.debug('Saved %d collections to %s' % (len(bag_cols), bagcolsfile))
                with open(tmpdir + SEP + 'VERSION', 'w') as (control):
                    control.write(APP['version'])
                target = self.srvutl.zip(export_path, tmpdir)
            return target
        except Exception as error:
            self.log.error('Errors raised during export: %s', error)
            self.log.error(self.get_traceback())
            return False

    def export_to_excel(self, bag, export_path=None):

        def as_text(value):
            if value is None:
                return ''
            else:
                return str(value)

        if len(bag) == 0:
            return
        try:
            wb = Workbook()
            highlight = NamedStyle(name='highlight')
            highlight.font = Font(name='DejaVu Sans', bold=True, size=10)
            bd = Side(style='thick', color='000000')
            highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)
            wb.add_named_style(highlight)
            normal = NamedStyle(name='normal')
            normal.font = Font(name='DejaVu Sans', bold=False, size=10)
            wb.add_named_style(normal)
            ws = wb.active
            header = []
            for field in HEADER_FIELDS:
                header.append(str(HEADER[field]))

            ws.append(header)
            data = []
            for sid in bag:
                row = []
                metadata = self.srvdtb.get_sapnote_metadata(sid)
                for field in HEADER_FIELDS:
                    if field == 'collections':
                        cols = ', '.join([self.srvclt.get_name_by_cid(col) for col in metadata[field]])
                        row.append(cols)
                    else:
                        if field == 'releasedon':
                            excel_date = self.srvutl.get_excel_date(metadata[field])
                            row.append(excel_date)
                        else:
                            row.append(str(metadata[field]))

                ws.append(row)
                data.append(row)

            for col in ('A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J'):
                cell = col + '1'
                ws[cell].style = 'highlight'

            for col in ('A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J'):
                for i in range(len(bag)):
                    cell = col + '%d' % (i + 2)
                    ws[cell].style = 'normal'

            try:
                for column_cells in ws.columns:
                    length = max(len(as_text(cell.value)) for cell in column_cells)
                    ws.column_dimensions[COL[column_cells[0].column]].width = length

            except Exception as error:
                self.log.error(error)
                self.log.error(self.get_traceback())
                self.log.error("This piece of code isn't working on Windows...")

            ws.auto_filter.ref = 'A1:J1'
            ws.auto_filter.add_sort_condition('I2:I%d' % len(bag))
            wb.save(export_path)
            return True
        except Exception as error:
            self.log.error(error)
            self.log.error(self.get_traceback())
            return False